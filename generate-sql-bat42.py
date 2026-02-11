#!/usr/bin/env python3
"""Generate SQL statements to import Lot 4.2 buildings from XLSX."""
import openpyxl

wb = openpyxl.load_workbook("/home/ubuntu/crm-e2mt2/batiments-4.2.xlsx")
ws = wb[wb.sheetnames[0]]

portfolio_map = {
    "FERROVIAIRE": "Ferroviaire",
    "FERROVIAIRE - PERIMETRE GARE": "Ferroviaire",
    "INDUSTRIEL": "Industriel",
    "GARES": "Gares",
    "TERTIAIRE": "Tertiaire",
    "SOCIAL": "Social",
}

def escape_sql(s):
    if s is None:
        return "NULL"
    s = str(s).replace("'", "''").replace("\\", "\\\\")
    return f"'{s}'"

buildings = []

for row_idx, row in enumerate(ws.iter_rows(min_row=10, values_only=True), 10):
    if not any(cell is not None for cell in row[:20]):
        continue
    
    integrated = str(row[46]).strip().upper() if row[46] else ""
    if integrated != "O":
        continue
    
    dept_num = str(row[4]).strip() if row[4] else ""
    dept_name = str(row[5]).strip() if row[5] else ""
    code_ut = str(row[7]).strip() if row[7] else ""
    nom_ut = str(row[8]).strip() if row[8] else ""
    num_bat = str(row[10]).strip() if row[10] else ""
    ut_bat = str(row[11]).strip() if row[11] else ""
    nom_immosis = str(row[12]).strip() if row[12] else ""
    nom_usuel = str(row[13]).strip() if row[13] else ""
    type_bat_marche = str(row[15]).strip() if row[15] else ""
    adresse = str(row[16]).strip() if row[16] else ""
    code_postal = str(row[17]).strip() if row[17] else ""
    ville = str(row[18]).strip() if row[18] else ""
    surface = row[19] if row[19] else None
    etat_emploi = str(row[20]).strip() if row[20] else ""
    portefeuille_raw = str(row[25]).strip().upper() if row[25] else ""
    proprietaire = str(row[23]).strip() if row[23] else ""
    occupant = str(row[26]).strip() if row[26] else ""
    
    name = nom_usuel if nom_usuel and nom_usuel != "None" and nom_usuel != nom_immosis else nom_immosis
    if not name or name == "None":
        name = f"Batiment {num_bat}"
    
    code = ut_bat if ut_bat and ut_bat != "None" else f"{code_ut}-{num_bat}"
    
    portfolio = "Ferroviaire"
    for key, val in portfolio_map.items():
        if key in portefeuille_raw:
            portfolio = val
            break
    
    parts = []
    if adresse and adresse != "None":
        parts.append(adresse)
    if code_postal and code_postal != "None":
        parts.append(code_postal)
    if ville and ville != "None":
        parts.append(ville)
    full_address = ", ".join(parts) if parts else None
    
    desc_parts = []
    if dept_name and dept_name != "None":
        desc_parts.append(f"Departement: {dept_name} ({dept_num})")
    if type_bat_marche and type_bat_marche != "None":
        desc_parts.append(f"Type: {type_bat_marche}")
    if proprietaire and proprietaire != "None":
        desc_parts.append(f"Proprietaire: {proprietaire}")
    if occupant and occupant != "None":
        desc_parts.append(f"Occupant: {occupant}")
    if etat_emploi and etat_emploi != "None":
        desc_parts.append(f"Etat: {etat_emploi}")
    if nom_ut and nom_ut != "None":
        desc_parts.append(f"UT: {nom_ut}")
    description = " | ".join(desc_parts) if desc_parts else None
    
    surface_val = "NULL"
    if surface is not None:
        try:
            surface_val = str(float(surface))
        except (ValueError, TypeError):
            surface_val = "NULL"
    
    buildings.append((name, code, portfolio, full_address, surface_val, description))

# Write SQL file
with open("/home/ubuntu/crm-e2mt2/import-bat42.sql", "w") as f:
    # First batch: delete and insert lot
    f.write("-- Batch 1: Cleanup\n")
    
    # Write insert statements in batches of 50
    batch_size = 50
    for i in range(0, len(buildings), batch_size):
        batch = buildings[i:i+batch_size]
        f.write(f"\n-- Batch insert {i+1}-{i+len(batch)}\n")
        values_list = []
        for name, code, portfolio, address, surface_val, description in batch:
            values_list.append(
                f"({escape_sql(name)}, {escape_sql(code)}, (SELECT id FROM lots WHERE code = '4.2'), "
                f"{escape_sql(portfolio)}, {escape_sql(address)}, {surface_val}, {escape_sql(description)}, 1)"
            )
        f.write(f"INSERT INTO buildings (name, code, lotId, portfolio, address, surface, description, isActive) VALUES\n")
        f.write(",\n".join(values_list))
        f.write(";\n")

print(f"Generated SQL for {len(buildings)} buildings")
print(f"SQL file: /home/ubuntu/crm-e2mt2/import-bat42.sql")
