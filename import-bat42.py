#!/usr/bin/env python3
"""Import Lot 4.2 (PACA) buildings from XLSX, replacing Lot 4.1 buildings."""
import os
import mysql.connector
import openpyxl

DATABASE_URL = os.environ.get("DATABASE_URL", "")

# Parse DATABASE_URL
url = DATABASE_URL.replace("mysql://", "")
if "?" in url:
    url = url.split("?")[0]
userpass, hostdb = url.split("@")
user, password = userpass.split(":")
hostport, dbname = hostdb.split("/")
host, port = hostport.split(":")

conn = mysql.connector.connect(
    host=host, port=int(port), user=user, password=password,
    database=dbname, ssl_disabled=False,
    ssl_ca="/etc/ssl/certs/ca-certificates.crt"
)
cursor = conn.cursor(dictionary=True)

# Step 1: Get or create Lot 4.2
cursor.execute("SELECT id FROM lots WHERE code = '4.2'")
lot_row = cursor.fetchone()
if lot_row:
    lot_id = lot_row["id"]
    print(f"Found existing Lot 4.2 with id={lot_id}")
else:
    cursor.execute("INSERT INTO lots (code, name, region) VALUES ('4.2', 'Lot 4.2', 'Provence-Alpes-Côte d''Azur')")
    lot_id = cursor.lastrowid
    print(f"Created Lot 4.2 with id={lot_id}")
    conn.commit()

# Step 2: Delete existing buildings (and their interventions/comments/history/alerts)
cursor.execute("SELECT id FROM buildings")
building_ids = [r["id"] for r in cursor.fetchall()]
if building_ids:
    ids_str = ",".join(str(bid) for bid in building_ids)
    # Delete related data first (foreign keys)
    cursor.execute(f"DELETE FROM intervention_comments WHERE interventionId IN (SELECT id FROM interventions WHERE buildingId IN ({ids_str}))")
    cursor.execute(f"DELETE FROM intervention_history WHERE interventionId IN (SELECT id FROM interventions WHERE buildingId IN ({ids_str}))")
    cursor.execute(f"DELETE FROM alerts WHERE interventionId IN (SELECT id FROM interventions WHERE buildingId IN ({ids_str}))")
    cursor.execute(f"DELETE FROM intervention_bpu_lines WHERE interventionId IN (SELECT id FROM interventions WHERE buildingId IN ({ids_str}))")
    cursor.execute(f"DELETE FROM interventions WHERE buildingId IN ({ids_str})")
    cursor.execute("DELETE FROM buildings")
    conn.commit()
    print(f"Deleted {len(building_ids)} existing buildings and their related data")

# Step 3: Read XLSX
wb = openpyxl.load_workbook("/home/ubuntu/crm-e2mt2/batiments-4.2.xlsx")
ws = wb[wb.sheetnames[0]]

# Portfolio mapping
portfolio_map = {
    "FERROVIAIRE": "Ferroviaire",
    "FERROVIAIRE - PERIMETRE GARE": "Ferroviaire",
    "INDUSTRIEL": "Industriel",
    "GARES": "Gares",
    "TERTIAIRE": "Tertiaire",
    "SOCIAL": "Social",
}

inserted = 0
skipped = 0

for row_idx, row in enumerate(ws.iter_rows(min_row=10, values_only=True), 10):
    # Skip empty rows
    if not any(cell is not None for cell in row[:20]):
        continue
    
    # Check if integrated in E2MT perimeter (col AV = index 47, value 'O')
    integrated = str(row[47]).strip().upper() if row[47] else ""
    if integrated != "O":
        skipped += 1
        continue
    
    # Extract fields (0-indexed from the tuple)
    lot_geo = str(row[1]).strip() if row[1] else ""
    region_admin = str(row[2]).strip() if row[2] else ""
    region_sncf = str(row[3]).strip() if row[3] else ""
    dept_num = str(row[4]).strip() if row[4] else ""
    dept_name = str(row[5]).strip() if row[5] else ""
    code_ut = str(row[7]).strip() if row[7] else ""
    nom_ut = str(row[8]).strip() if row[8] else ""
    num_bat = str(row[10]).strip() if row[10] else ""
    ut_bat = str(row[11]).strip() if row[11] else ""
    nom_immosis = str(row[12]).strip() if row[12] else ""
    nom_usuel = str(row[13]).strip() if row[13] else ""
    type_bat_immosis = str(row[14]).strip() if row[14] else ""
    type_bat_marche = str(row[15]).strip() if row[15] else ""
    adresse = str(row[16]).strip() if row[16] else ""
    code_postal = str(row[17]).strip() if row[17] else ""
    ville = str(row[18]).strip() if row[18] else ""
    surface = row[19] if row[19] else None
    etat_emploi = str(row[20]).strip() if row[20] else ""
    portefeuille_raw = str(row[25]).strip().upper() if row[25] else ""
    proprietaire = str(row[23]).strip() if row[23] else ""
    occupant = str(row[26]).strip() if row[26] else ""
    
    # Build name (prefer usuel, fallback to immosis)
    name = nom_usuel if nom_usuel and nom_usuel != "None" and nom_usuel != nom_immosis else nom_immosis
    if not name or name == "None":
        name = f"Bâtiment {num_bat}"
    
    # Build code
    code = ut_bat if ut_bat and ut_bat != "None" else f"{code_ut}-{num_bat}"
    
    # Map portfolio
    portfolio = "Ferroviaire"  # default
    for key, val in portfolio_map.items():
        if key in portefeuille_raw:
            portfolio = val
            break
    
    # Build address
    full_address = ""
    parts = []
    if adresse and adresse != "None":
        parts.append(adresse)
    if code_postal and code_postal != "None":
        parts.append(code_postal)
    if ville and ville != "None":
        parts.append(ville)
    full_address = ", ".join(parts) if parts else None
    
    # Build description
    desc_parts = []
    if dept_name and dept_name != "None":
        desc_parts.append(f"Département: {dept_name} ({dept_num})")
    if type_bat_marche and type_bat_marche != "None":
        desc_parts.append(f"Type: {type_bat_marche}")
    if proprietaire and proprietaire != "None":
        desc_parts.append(f"Propriétaire: {proprietaire}")
    if occupant and occupant != "None":
        desc_parts.append(f"Occupant: {occupant}")
    if etat_emploi and etat_emploi != "None":
        desc_parts.append(f"État: {etat_emploi}")
    if nom_ut and nom_ut != "None":
        desc_parts.append(f"UT: {nom_ut}")
    description = " | ".join(desc_parts) if desc_parts else None
    
    # Clean surface
    surface_val = None
    if surface is not None:
        try:
            surface_val = float(surface)
        except (ValueError, TypeError):
            surface_val = None
    
    try:
        cursor.execute(
            """INSERT INTO buildings (name, code, lotId, portfolio, address, surface, description, isActive)
               VALUES (%s, %s, %s, %s, %s, %s, %s, 1)""",
            (name, code, lot_id, portfolio, full_address, surface_val, description)
        )
        inserted += 1
    except Exception as e:
        print(f"  Error row {row_idx}: {e} - name={name}, code={code}")
        continue

conn.commit()
print(f"\nImport complete!")
print(f"Inserted: {inserted} buildings")
print(f"Skipped (not in E2MT perimeter): {skipped}")

# Verify
cursor.execute("SELECT COUNT(*) as cnt FROM buildings WHERE lotId = %s", (lot_id,))
print(f"Total buildings in DB for Lot 4.2: {cursor.fetchone()['cnt']}")

cursor.close()
conn.close()
