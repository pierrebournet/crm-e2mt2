#!/usr/bin/env python3
"""Debug: check integration column values."""
import openpyxl

wb = openpyxl.load_workbook("/home/ubuntu/crm-e2mt2/batiments-4.2.xlsx")
ws = wb[wb.sheetnames[0]]

# Check row 9 header for integration column
row9 = list(ws.iter_rows(min_row=9, max_row=9, values_only=True))[0]
for i, cell in enumerate(row9):
    if cell and "INTEGRA" in str(cell).upper():
        print(f"Col {i}: {cell}")
    if cell and ("O/N" in str(cell) or "PERIMETRE" in str(cell).upper()):
        print(f"Col {i}: {cell}")

# Check first 5 data rows for all columns around 45-55
print("\n--- Data rows 10-14, columns 45-55 ---")
for row_idx, row in enumerate(ws.iter_rows(min_row=10, max_row=14, values_only=True), 10):
    vals = {i: row[i] for i in range(45, min(55, len(row))) if row[i] is not None}
    print(f"Row {row_idx}: {vals}")

# Also check columns around the O/N
print("\n--- Row 8 headers around col 45-55 ---")
row8 = list(ws.iter_rows(min_row=8, max_row=8, values_only=True))[0]
for i in range(45, min(55, len(row8))):
    if row8[i]:
        print(f"Col {i}: {row8[i]}")

# Check total rows with any data
count = 0
for row in ws.iter_rows(min_row=10, values_only=True):
    if row[11] and str(row[11]).strip() != "None":  # UT-BAT column
        count += 1
print(f"\nTotal rows with UT-BAT: {count}")
