#!/usr/bin/env python3
"""Read the Lot 4.2 buildings XLSX file to understand its structure."""
import openpyxl

wb = openpyxl.load_workbook("/home/ubuntu/crm-e2mt2/batiments-4.2.xlsx")
print(f"Sheets: {wb.sheetnames}")

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"\n=== Sheet: {sheet_name} ===")
    print(f"Dimensions: {ws.dimensions}")
    print(f"Max rows: {ws.max_row}, Max cols: {ws.max_column}")
    
    # Print first 10 rows
    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        if i <= 10:
            print(f"Row {i}: {row}")
        else:
            break
    
    # Count non-empty rows
    count = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if any(cell is not None for cell in row):
            count += 1
    print(f"\nTotal data rows: {count}")
