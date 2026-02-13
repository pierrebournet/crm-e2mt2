import openpyxl

wb = openpyxl.load_workbook('/home/ubuntu/upload/NoticeAT-nommageettype.xlsx')
ws = wb.active

print(f"Rows: {ws.max_row}, Cols: {ws.max_column}")
print("=" * 120)

for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=False):
    values = [str(cell.value) if cell.value is not None else "" for cell in row]
    print(f"Row {row[0].row:3d}: | {'  |  '.join(values)}")
